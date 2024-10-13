# Step 1: Design the Code
# =========================
# 
# The project will be structured as follows:
# 
# 1. **Collecting User Input**: Start with a command-line prompt to select a CSV file.
# 2. **CSV Data Handling**: Load the CSV using pandas for easier manipulation.
# 3. **Excel Formatting**: Use openpyxl to write to Excel and apply formatting, such as bold headers and colors.
# 4. **Optional Enhancements**: 
#     - Ability to split data into different sheets based on unique values in a specified column.
#     - Optionally apply filters automatically to each sheet.
# 5. **User Interface**: Begin as a CLI tool, with a later option to create a graphical interface using Tkinter.

# Step 2: Write the Code
# =======================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from tkinter import Tk, filedialog
import os

def collect_csv_file():
    root = Tk()
    root.withdraw()  # Hides the root window
    file_path = filedialog.askopenfilename(
        title="Select CSV File",
        filetypes=[("CSV files", "*.csv")]
    )
    return file_path

def format_excel(ws):
    # Formatting the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill

    # Auto-adjust column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

def generate_excel(csv_file, output_file, split_by_column=None):
    df = pd.read_csv(csv_file)
    wb = Workbook()
    
    if split_by_column and split_by_column in df.columns:
        unique_values = df[split_by_column].unique()
        for value in unique_values:
            ws = wb.create_sheet(title=str(value))
            value_df = df[df[split_by_column] == value]
            for r_idx, row in enumerate(value_df.itertuples(), start=1):
                for c_idx, value in enumerate(row[1:], start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            format_excel(ws)
        wb.remove(wb["Sheet"])  # Remove the default sheet
    else:
        ws = wb.active
        ws.title = "Data"
        for r_idx, row in enumerate(df.itertuples(), start=1):
            for c_idx, value in enumerate(row[1:], start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        format_excel(ws)

    wb.save(output_file)
    print(f"Excel file saved to: {output_file}")

def main():
    print("CSV to Excel Formatter Tool")
    csv_file = collect_csv_file()
    if not csv_file or not os.path.exists(csv_file):
        print("Invalid file. Please try again.")
        return

    output_file = csv_file.replace('.csv', '_formatted.xlsx')
    split_column = input("Enter column name to split data into different sheets (optional): ")
    
    generate_excel(csv_file, output_file, split_by_column=split_column if split_column else None)
    print("Formatting complete.")

if __name__ == "__main__":
    main()
