"""CSV to formatted Excel conversion tool."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import os
import sys

def select_csv_file_dialog():
    """Opens a Tkinter dialog to select a CSV file."""
    root = Tk()
    root.withdraw()  
    root.attributes('-topmost', True)  
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo CSV",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    root.destroy() 
    if not file_path:
        print("Nenhum arquivo selecionado. Saindo.")
        sys.exit(0) 
    return file_path

def apply_excel_formatting(worksheet):
    """Applies header formatting and auto-adjusts column widths."""
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    

    for cell in worksheet[1]: 
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, column in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value:
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
            except:
                pass 
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_excel_from_csv(csv_file_path, output_excel_path, split_column_name=None):
    """
    Reads a CSV, optionally splits data by a column into sheets,
    formats, and saves as an Excel file.
    """
    try:
        
        if not csv_file_path.lower().endswith(".csv"):
            print(f"Erro: O arquivo '{csv_file_path}' não parece ser um CSV.")
            return False
        df = pd.read_csv(csv_file_path)
    except FileNotFoundError:
        print(f"Erro: Arquivo CSV não encontrado em '{csv_file_path}'.")
        return False
    except pd.errors.EmptyDataError:
        print(f"Erro: O arquivo CSV '{csv_file_path}' está vazio.")
        return False
    except pd.errors.ParserError:
        print(f"Erro: Não foi possível parsear o arquivo CSV '{csv_file_path}'. Verifique o formato.")
        return False
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao ler o CSV: {e}")
        return False

    workbook = Workbook()
    default_sheet_removed = False

    if split_column_name and split_column_name in df.columns:
        return
        
        def sanitize_sheet_name(name):
            invalid_chars = ['*', ':', '/', '\\', '?', '[', ']']
            for char in invalid_chars:
                name = name.replace(char, '_')
            return name[:31] 

        unique_values = df[split_column_name].unique()
        for value in unique_values:
            sheet_name = sanitize_sheet_name(str(value))
            if sheet_name in workbook.sheetnames:
                sheet_name = f"{sheet_name}_{pd.util.hash_pandas_object(pd.Series(value)).sum() % 1000}" # Adiciona um hash para unicidade
            
            value_df = df[df[split_column_name] == value]
            
            if not value_df.empty:
                worksheet = workbook.create_sheet(title=sheet_name)
                for r_idx, row in enumerate(dataframe_to_rows(value_df, index=False, header=True), 1):
                    for c_idx, cell_value in enumerate(row, 1):
                        worksheet.cell(row=r_idx, column=c_idx, value=cell_value)
                apply_excel_formatting(worksheet)

        
        if workbook.sheetnames != ["Sheet"] and "Sheet" in workbook.sheetnames:
             del workbook["Sheet"]
             default_sheet_removed = True
    else:
        if "Sheet" in workbook.sheetnames:
            worksheet = workbook["Sheet"]
            worksheet.title = "Dados"
        else:
            worksheet = workbook.create_sheet(title="Dados")
        if df.empty:
            print("Warning: O CSV está vazio, a planilha 'Dados' será criada sem dados.")
            if not df.columns.empty: 
                 for c_idx, col_name in enumerate(df.columns, 1):
                     worksheet.cell(row=1, column=c_idx, value=col_name)
            apply_excel_formatting(worksheet)
        else:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, cell_value in enumerate(row, 1):
                    worksheet.cell(row=r_idx, column=c_idx, value=cell_value)
            apply_excel_formatting(worksheet)
    
  
    if not default_sheet_removed and len(workbook.sheetnames) > 1 and "Sheet" in workbook.sheetnames and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1 and workbook["Sheet"].cell(1,1).value is None:
      
        try:
            del workbook["Sheet"]
        except KeyError:
            pass 

    try:
        workbook.save(output_excel_path)
        print(f"Arquivo Excel salvo com sucesso em: {output_excel_path}")
        return True
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")
        print("Verifique se o arquivo não está aberto em outro programa e se você tem permissão para escrever no local.")
        return False

def main():
    print("--- Ferramenta de Conversão CSV para Excel Formatado ---")
    
    csv_file = select_csv_file_dialog()
    if not os.path.isfile(csv_file):
        print(f"Erro: O caminho '{csv_file}' não é um arquivo válido.")
        return

    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_dir = os.path.dirname(csv_file)
    default_output_file = os.path.join(output_dir, f"{base_name}_formatado.xlsx")
    
    output_file_input = input(f"Salvar como [{default_output_file}]: ") or default_output_file
    output_excel_path = os.path.abspath(output_file_input) 

    split_column = input("Digite o nome da coluna para dividir em abas (opcional, pressione Enter para pular): ").strip()
    
    print(f"\nProcessando '{csv_file}'...")
    success = generate_excel_from_csv(
        csv_file,
        output_excel_path,
        split_by_column=split_column if split_column else None
    )
    
    if success:
        print("Conversão e formatação concluídas.")
    else:
        print("A conversão falhou. Veja as mensagens de erro acima.")

if __name__ == "__main__":
    main()
